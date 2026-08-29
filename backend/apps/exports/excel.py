import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .donnees import rassembler_donnees_patient

STYLE_ENTETE_FONT = Font(bold=True, color="FFFFFF")
STYLE_ENTETE_FOND = PatternFill(start_color="2C5F8A", end_color="2C5F8A", fill_type="solid")


def _ecrire_feuille(classeur, nom, entetes, lignes):
    feuille = classeur.create_sheet(nom)
    feuille.append(entetes)
    for cellule in feuille[1]:
        cellule.font = STYLE_ENTETE_FONT
        cellule.fill = STYLE_ENTETE_FOND
    for ligne in lignes:
        feuille.append(ligne)
    for index, entete in enumerate(entetes, start=1):
        feuille.column_dimensions[get_column_letter(index)].width = max(15, len(str(entete)) + 4)
    return feuille


def generer_excel_patient(patient) -> bytes:
    donnees = rassembler_donnees_patient(patient)
    classeur = Workbook()
    classeur.remove(classeur.active)  # on retire la feuille vide par défaut

    _ecrire_feuille(
        classeur, "Prescriptions actives",
        ["Médicament", "Type", "Dose", "Prescripteur", "Depuis le", "Instructions"],
        [
            [
                p.medicament.denomination, p.get_type_prise_display(),
                f"{p.dose_quantite} {p.dose_unite}", p.medecin_prescripteur.get_full_name(),
                p.date_debut.strftime("%d/%m/%Y"), p.instructions,
            ]
            for p in donnees["prescriptions_actives"]
        ],
    )

    _ecrire_feuille(
        classeur, "Interactions",
        ["Substance A", "Médicament A", "Substance B", "Médicament B", "Niveau", "Détail"],
        [
            [i.substance_a, i.medicament_a, i.substance_b, i.medicament_b, i.niveau, i.libelle]
            for i in donnees["interactions"]
        ],
    )

    _ecrire_feuille(
        classeur, "Historique des prises",
        ["Date/heure", "Médicament", "Quantité", "Statut"],
        [
            [
                prise.date_heure_reelle.strftime("%d/%m/%Y %H:%M") if prise.date_heure_reelle else "",
                prise.prescription.medicament.denomination,
                float(prise.quantite_prise) if prise.quantite_prise is not None else None,
                prise.get_statut_display(),
            ]
            for prise in donnees["historique_prises"]
        ],
    )

    _ecrire_feuille(
        classeur, "Stock",
        ["Médicament", "Quantité restante", "Quantité initiale", "Péremption", "En alerte"],
        [
            [
                boite.medicament.denomination, float(boite.quantite_restante), float(boite.quantite_initiale),
                boite.date_peremption.strftime("%d/%m/%Y") if boite.date_peremption else "",
                "Oui" if boite.en_alerte else "Non",
            ]
            for boite in donnees["boites_actives"]
        ],
    )

    tampon = io.BytesIO()
    classeur.save(tampon)
    return tampon.getvalue()
