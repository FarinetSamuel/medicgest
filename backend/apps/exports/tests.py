import datetime
import io

from django.contrib.auth.models import Group
from django.test import TestCase
from openpyxl import load_workbook
from rest_framework.test import APITestCase

from apps.medicaments.models import Medicament
from apps.patients.models import Patient, PatientMedecin
from apps.prescriptions.models import Prescription, Prise
from apps.utilisateurs.models import ROLE_MEDECIN, ROLE_PATIENT, Utilisateur

from .donnees import rassembler_donnees_patient
from .excel import generer_excel_patient
from .pdf import generer_pdf_patient


def creer_utilisateur_avec_role(email, role):
    user = Utilisateur.objects.create_user(username=email.split("@")[0], email=email, password="x")
    groupe, _ = Group.objects.get_or_create(name=role)
    user.groups.add(groupe)
    return user


class DonneesExportTest(TestCase):
    def setUp(self):
        self.medecin = creer_utilisateur_avec_role("medexp1@example.com", ROLE_MEDECIN)
        user_patient = creer_utilisateur_avec_role("patexp1@example.com", ROLE_PATIENT)
        self.patient = Patient.objects.create(
            utilisateur=user_patient,
            numero_dossier="DOS-EXP-1",
            date_naissance=datetime.date(1980, 1, 1),
            sexe=Patient.Sexe.FEMININ,
        )
        self.medicament = Medicament.objects.create(code_cis="EXP1", denomination="EXPOL")
        self.prescription = Prescription.objects.create(
            patient=self.patient,
            medicament=self.medicament,
            medecin_prescripteur=self.medecin,
            type_prise=Prescription.TypePrise.RESERVE,
            dose_quantite=1,
            dose_unite="comprimé",
            date_debut=datetime.date(2026, 1, 1),
            statut=Prescription.Statut.ACTIVE,
        )

    def test_rassemble_les_prescriptions_actives(self):
        donnees = rassembler_donnees_patient(self.patient)
        self.assertEqual(len(donnees["prescriptions_actives"]), 1)

    def test_prescription_arretee_absente(self):
        self.prescription.statut = Prescription.Statut.ARRETEE
        self.prescription.save()
        donnees = rassembler_donnees_patient(self.patient)
        self.assertEqual(len(donnees["prescriptions_actives"]), 0)


class GenerationPdfTest(TestCase):
    def setUp(self):
        self.medecin = creer_utilisateur_avec_role("medexp2@example.com", ROLE_MEDECIN)
        user_patient = creer_utilisateur_avec_role("patexp2@example.com", ROLE_PATIENT)
        self.patient = Patient.objects.create(
            utilisateur=user_patient,
            numero_dossier="DOS-EXP-2",
            date_naissance=datetime.date(1980, 1, 1),
            sexe=Patient.Sexe.MASCULIN,
        )

    def test_genere_un_vrai_pdf_valide(self):
        """Vérifie la signature binaire réelle d'un PDF (%PDF-), pas juste l'absence d'exception."""
        contenu = generer_pdf_patient(self.patient)
        self.assertTrue(contenu.startswith(b"%PDF-"))
        self.assertGreater(len(contenu), 500)

    def test_pdf_genere_meme_sans_aucune_donnee(self):
        """Un patient sans prescription/prise/stock ne doit pas faire planter la génération."""
        contenu = generer_pdf_patient(self.patient)
        self.assertTrue(contenu.startswith(b"%PDF-"))


class GenerationExcelTest(TestCase):
    def setUp(self):
        self.medecin = creer_utilisateur_avec_role("medexp3@example.com", ROLE_MEDECIN)
        user_patient = creer_utilisateur_avec_role("patexp3@example.com", ROLE_PATIENT)
        self.patient = Patient.objects.create(
            utilisateur=user_patient,
            numero_dossier="DOS-EXP-3",
            date_naissance=datetime.date(1980, 1, 1),
            sexe=Patient.Sexe.FEMININ,
        )
        self.medicament = Medicament.objects.create(code_cis="EXP3", denomination="EXPOL3")
        self.prescription = Prescription.objects.create(
            patient=self.patient,
            medicament=self.medicament,
            medecin_prescripteur=self.medecin,
            type_prise=Prescription.TypePrise.RESERVE,
            dose_quantite=1,
            dose_unite="comprimé",
            date_debut=datetime.date(2026, 1, 1),
            statut=Prescription.Statut.ACTIVE,
        )

    def test_genere_un_vrai_classeur_avec_les_bonnes_feuilles(self):
        """Réouvre le fichier généré avec openpyxl pour valider son contenu réel, pas juste sa génération."""
        contenu = generer_excel_patient(self.patient)
        classeur = load_workbook(io.BytesIO(contenu))

        self.assertEqual(
            set(classeur.sheetnames),
            {"Prescriptions actives", "Interactions", "Historique des prises", "Stock"},
        )

    def test_feuille_prescriptions_contient_la_bonne_ligne(self):
        contenu = generer_excel_patient(self.patient)
        classeur = load_workbook(io.BytesIO(contenu))
        feuille = classeur["Prescriptions actives"]

        self.assertEqual(feuille.cell(row=1, column=1).value, "Médicament")
        self.assertEqual(feuille.cell(row=2, column=1).value, "EXPOL3")


class ExportAPITest(APITestCase):
    def setUp(self):
        self.medecin_suiveur = creer_utilisateur_avec_role("medexpapi1@example.com", ROLE_MEDECIN)
        self.medecin_autre = creer_utilisateur_avec_role("medexpapi2@example.com", ROLE_MEDECIN)
        self.user_patient = creer_utilisateur_avec_role("patexpapi1@example.com", ROLE_PATIENT)
        self.patient = Patient.objects.create(
            utilisateur=self.user_patient,
            numero_dossier="DOS-EXP-API-1",
            date_naissance=datetime.date(1980, 1, 1),
            sexe=Patient.Sexe.MASCULIN,
        )
        PatientMedecin.objects.create(patient=self.patient, medecin=self.medecin_suiveur, actif=True)

    def test_medecin_suiveur_peut_telecharger_le_pdf(self):
        self.client.force_authenticate(self.medecin_suiveur)
        response = self.client.get(f"/api/v1/patients/{self.patient.id}/export-pdf/")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertIn("attachment", response["Content-Disposition"])

    def test_medecin_suiveur_peut_telecharger_l_excel(self):
        self.client.force_authenticate(self.medecin_suiveur)
        response = self.client.get(f"/api/v1/patients/{self.patient.id}/export-excel/")
        self.assertEqual(response.status_code, 200)
        self.assertIn("spreadsheetml", response["Content-Type"])

    def test_medecin_non_suiveur_ne_peut_pas_exporter(self):
        self.client.force_authenticate(self.medecin_autre)
        response = self.client.get(f"/api/v1/patients/{self.patient.id}/export-pdf/")
        self.assertEqual(response.status_code, 404)

    def test_patient_peut_exporter_ses_propres_donnees(self):
        self.client.force_authenticate(self.user_patient)
        response = self.client.get(f"/api/v1/patients/{self.patient.id}/export-excel/")
        self.assertEqual(response.status_code, 200)

    def test_anonyme_refuse(self):
        response = self.client.get(f"/api/v1/patients/{self.patient.id}/export-pdf/")
        self.assertEqual(response.status_code, 403)
